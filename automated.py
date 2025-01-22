
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from datetime import datetime
import time

today = datetime.today().strftime('%A')

excel_path = r"keywords.xlsx"
workbook = load_workbook(excel_path)

if today not in workbook.sheetnames:
    print(f"Error: Sheet for '{today}' not found in the Excel file.")
    exit()

sheet = workbook[today]

chromedriver_path = r"C:\Windows\chromedriver.exe"
service = Service(chromedriver_path)
driver = webdriver.Chrome(service=service)

for row in range(2, sheet.max_row + 1):  
    keyword = sheet[f"C{row}"].value
    if not keyword:
        continue 

    driver.get("https://www.google.com") 
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(keyword)
    search_box.send_keys(Keys.RETURN)
    time.sleep(8)

    results = driver.find_elements(By.CSS_SELECTOR, "h3")
    options = [result.text for result in results if result.text]

    if options:
        longest_option = max(options, key=len)  
        shortest_option = min(options, key=len)

        sheet[f"D{row}"] = longest_option
        sheet[f"E{row}"] = shortest_option
    else:
        sheet[f"D{row}"] = "No results"
        sheet[f"E{row}"] = "No results"

workbook.save(excel_path)
print(f"Processing completed. Results saved in '{excel_path}'.")

driver.quit()
